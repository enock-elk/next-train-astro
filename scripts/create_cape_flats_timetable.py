"""Create Cape Flats Line Public Holiday Timetable 2026 spreadsheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

out = Path(__file__).resolve().parents[1] / "Cape-Flats-Line-Public-Holiday-Timetable-2026.xlsx"

wb = Workbook()

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
station_fill = PatternFill("solid", fgColor="D6EAF8")
station_font = Font(bold=True, name="Calibri", size=11)
title_font = Font(bold=True, name="Calibri", size=14, color="1F4E79")
subtitle_font = Font(bold=True, name="Calibri", size=12, color="1F4E79")
cell_font = Font(name="Calibri", size=11)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
alt_fill = PatternFill("solid", fgColor="F5F9FC")

outbound_trains = [
    "0501", "0503", "0505", "0507", "0509", "0511",
    "0513", "0515", "0517", "0519", "0521", "0523",
]
outbound_stations = [
    "CAPE TOWN", "WOODSTOCK", "SALT RIVER", "KOEBERG RD", "MAITLAND", "NDABENI",
    "PINELANDS", "HAZENDAL", "ATHLONE", "CRAWFORD", "LANSDOWNE", "WETTON",
    "OTTERY", "SOUTHFIELD", "HEATHFIELD", "RETREAT",
]
outbound = [
    ["06:00", "07:10", "07:40", "08:50", "09:30", "10:30", "11:30", "12:30", "13:30", "14:10", "15:10", "16:00"],
    ["06:03", "07:13", "07:43", "08:53", "09:33", "10:33", "11:33", "12:33", "13:33", "14:13", "15:13", "16:03"],
    ["06:06", "07:16", "07:46", "08:56", "09:36", "10:36", "11:36", "12:36", "13:36", "14:16", "15:16", "16:06"],
    ["06:08", "07:18", "07:48", "08:58", "09:38", "10:38", "11:38", "12:38", "13:38", "14:18", "15:18", "16:08"],
    ["06:10", "07:20", "07:50", "09:00", "09:40", "10:40", "11:40", "12:40", "13:40", "14:20", "15:20", "16:10"],
    ["06:13", "07:23", "07:53", "09:03", "09:43", "10:43", "11:43", "12:43", "13:43", "14:23", "15:23", "16:13"],
    ["06:16", "07:26", "07:56", "09:06", "09:46", "10:46", "11:46", "12:46", "13:46", "14:26", "15:26", "16:16"],
    ["06:18", "07:28", "07:58", "09:08", "09:48", "10:48", "11:48", "12:48", "13:48", "14:28", "15:28", "16:18"],
    ["06:21", "07:31", "08:01", "09:11", "09:51", "10:51", "11:51", "12:51", "13:51", "14:31", "15:31", "16:21"],
    ["06:23", "07:33", "08:03", "09:13", "09:53", "10:53", "11:53", "12:53", "13:53", "14:33", "15:33", "16:23"],
    ["06:25", "07:35", "08:05", "09:15", "09:55", "10:55", "11:55", "12:55", "13:55", "14:35", "15:35", "16:25"],
    ["06:28", "07:38", "08:08", "09:18", "09:58", "10:58", "11:58", "12:58", "13:58", "14:38", "15:38", "16:28"],
    ["06:30", "07:40", "08:10", "09:20", "10:00", "11:00", "12:00", "13:00", "14:00", "14:40", "15:40", "16:30"],
    ["06:34", "07:44", "08:14", "09:24", "10:04", "11:04", "12:04", "13:04", "14:04", "14:44", "15:44", "16:34"],
    ["06:38", "07:48", "08:18", "09:28", "10:08", "11:08", "12:08", "13:08", "14:08", "14:48", "15:48", "16:38"],
    ["06:40", "07:50", "08:20", "09:30", "10:10", "11:10", "12:10", "13:10", "14:10", "14:50", "15:50", "16:40"],
]

inbound_trains = [
    "0500", "0502", "0504", "0506", "0508", "0510",
    "0512", "0514", "0516", "0518", "0520", "0522",
]
inbound_stations = [
    "RETREAT", "HEATHFIELD", "SOUTHFIELD", "OTTERY", "WETTON", "LANSDOWNE",
    "CRAWFORD", "ATHLONE", "HAZENDAL", "PINELANDS", "NDABENI", "MAITLAND",
    "KOEBERG RD", "SALT RIVER", "WOODSTOCK", "CAPE TOWN",
]
inbound = [
    ["06:20", "06:50", "08:00", "08:40", "09:40", "10:20", "11:20", "12:20", "13:20", "14:20", "15:10", "16:00"],
    ["06:23", "06:53", "08:03", "08:43", "09:43", "10:23", "11:23", "12:23", "13:23", "14:23", "15:13", "16:03"],
    ["06:26", "06:56", "08:06", "08:46", "09:46", "10:26", "11:26", "12:26", "13:26", "14:26", "15:16", "16:06"],
    ["06:29", "06:59", "08:09", "08:49", "09:49", "10:29", "11:29", "12:29", "13:29", "14:29", "15:19", "16:09"],
    ["06:32", "07:02", "08:12", "08:52", "09:52", "10:32", "11:32", "12:32", "13:32", "14:32", "15:22", "16:12"],
    ["06:34", "07:04", "08:14", "08:54", "09:54", "10:34", "11:34", "12:34", "13:34", "14:34", "15:24", "16:14"],
    ["06:36", "07:06", "08:16", "08:56", "09:56", "10:36", "11:36", "12:36", "13:36", "14:36", "15:26", "16:16"],
    ["06:39", "07:09", "08:19", "08:59", "09:59", "10:39", "11:39", "12:39", "13:39", "14:39", "15:29", "16:19"],
    ["06:41", "07:11", "08:21", "09:01", "10:01", "10:41", "11:41", "12:41", "13:41", "14:41", "15:31", "16:21"],
    ["06:44", "07:14", "08:24", "09:04", "10:04", "10:44", "11:44", "12:44", "13:44", "14:44", "15:34", "16:24"],
    ["06:46", "07:16", "08:26", "09:06", "10:06", "10:46", "11:46", "12:46", "13:46", "14:46", "15:36", "16:26"],
    ["06:49", "07:19", "08:29", "09:09", "10:09", "10:49", "11:49", "12:49", "13:49", "14:49", "15:39", "16:29"],
    ["06:51", "07:21", "08:31", "09:11", "10:11", "10:51", "11:51", "12:51", "13:51", "14:51", "15:41", "16:31"],
    ["06:53", "07:23", "08:33", "09:13", "10:13", "10:53", "11:53", "12:53", "13:53", "14:53", "15:43", "16:33"],
    ["06:57", "07:27", "08:37", "09:17", "10:17", "10:57", "11:57", "12:57", "13:57", "14:57", "15:47", "16:37"],
    ["07:00", "07:30", "08:40", "09:20", "10:20", "11:00", "12:00", "13:00", "14:00", "15:00", "15:50", "16:40"],
]


def write_sheet(ws, title, direction, trains, stations, rows):
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(trains) + 1)
    ws["A2"] = direction
    ws["A2"].font = subtitle_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(trains) + 1)
    ws["A3"] = "Source: Metrorail / PRASA Cape Flats Line Public Holiday Timetable 2026"
    ws["A3"].font = Font(italic=True, name="Calibri", size=10, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(trains) + 1)

    headers = ["STATION / TRAIN NO."] + trains
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for r_idx, station in enumerate(stations):
        row_num = 6 + r_idx
        sc = ws.cell(row=row_num, column=1, value=station)
        sc.font = station_font
        sc.fill = station_fill
        sc.alignment = left
        sc.border = thin
        for c_idx, val in enumerate(rows[r_idx], 2):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font = cell_font
            cell.alignment = center
            cell.border = thin
            if r_idx % 2 == 1:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 20
    for i in range(2, len(trains) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = "B6"
    ws.row_dimensions[5].height = 22


ws1 = wb.active
ws1.title = "Cape Town to Retreat"
write_sheet(
    ws1,
    "CAPE FLATS LINE - PUBLIC HOLIDAY TIMETABLE 2026",
    "Direction: Cape Town to Retreat",
    outbound_trains,
    outbound_stations,
    outbound,
)

ws2 = wb.create_sheet("Retreat to Cape Town")
write_sheet(
    ws2,
    "CAPE FLATS LINE - PUBLIC HOLIDAY TIMETABLE 2026",
    "Direction: Retreat to Cape Town",
    inbound_trains,
    inbound_stations,
    inbound,
)

ws3 = wb.create_sheet("All trips (long format)")
long_headers = ["Direction", "Train No.", "Station", "Station Order", "Time"]
for col, h in enumerate(long_headers, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin

row = 2
for t_idx, train in enumerate(outbound_trains):
    for s_idx, station in enumerate(outbound_stations):
        values = [
            "Cape Town to Retreat",
            train,
            station,
            s_idx + 1,
            outbound[s_idx][t_idx],
        ]
        for c, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=c, value=val)
            cell.border = thin
            cell.font = cell_font
            cell.alignment = center if c != 3 else left
        row += 1

for t_idx, train in enumerate(inbound_trains):
    for s_idx, station in enumerate(inbound_stations):
        values = [
            "Retreat to Cape Town",
            train,
            station,
            s_idx + 1,
            inbound[s_idx][t_idx],
        ]
        for c, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=c, value=val)
            cell.border = thin
            cell.font = cell_font
            cell.alignment = center if c != 3 else left
        row += 1

ws3.column_dimensions["A"].width = 24
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 10
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:E{row - 1}"

wb.save(out)
print(f"Saved: {out}")
print(f"Exists: {out.exists()} Size: {out.stat().st_size} bytes")
