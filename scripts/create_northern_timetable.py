from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timedelta

out = Path(r"C:\Users\enock\OneDrive\Documents\GitHub\Metrorail Next Train\Source Code\Northern-Line-Public-Holiday-Timetable-2026.xlsx")

def tadd(base, mins):
    hh, mm = map(int, base.split(":"))
    dt = datetime(2000, 1, 1, hh, mm) + timedelta(minutes=mins)
    return dt.strftime("%H:%M")

E = "--"  # as printed on the timetable

ob_trains = ["2501","3201","2503","3203","2505","3205","2507","3207","2509","3209","2511","3211","2513","3213"]

# 25xx Cape Town departures
ct25 = {
    "2501": "06:15", "2503": "08:00", "2505": "09:40", "2507": "11:30",
    "2509": "12:55", "2511": "14:30", "2513": "16:10",
}
# offsets from Cape Town for 25xx
ct_offs = [
    ("CAPE TOWN", 0),
    ("WOODSTOCK", 3),
    ("SALT RIVER", 7),
    ("KOEBERG RD", 10),
    ("MAITLAND", 13),
    ("MUTUAL", 18),
    ("THORNTON", 21),
    ("GOODWOOD", 26),
    ("VASCO", 33),
    ("ELSIES RIVER", 36),
    ("PAROW", 40),
    ("TYGERBERG", 45),
    ("BELLVILLE", 50),  # arrive
]

# Bellville depart for all trains
bv_dep = {
    "2501": "07:06", "3201": "07:20", "2503": "08:51", "3203": "09:15",
    "2505": "10:31", "3205": "10:40", "2507": "12:21", "3207": "12:40",
    "2509": "13:46", "3209": "13:55", "2511": "15:21", "3211": "16:00",
    "2513": "17:01", "3213": "17:15",
}
platform_ob = {
    "2501": "4", "3201": "9", "2503": "4", "3203": "9",
    "2505": "4", "3205": "9", "2507": "4", "3207": "9",
    "2509": "4", "3209": "9", "2511": "4", "3211": "9",
    "2513": "4", "3213": "9",
}

# 25xx after Bellville -> Kraaifontein (offsets from Bellville Dep)
kf_offs = [
    ("STIKLAND", 9),
    ("BRACKENFELL", 17),
    ("EIKENFONTEIN", 22),
    ("KRAAIFONTEIN", 25),
]

# 32xx after Bellville -> Strand (offsets from Bellville Dep) verified against 3201
# 07:20 -> Kuils 07:32 (+12), BH 07:41 (+21), MR 07:49 (+29), ER 07:55 (+35),
# Faure 07:58 (+38), Fir 08:02 (+42), SW 08:14 (+54), VDS 08:17 (+57), Strand 08:20 (+60)
st_offs = [
    ("KUILS RIVER", 12),
    ("BLACKHEATH", 21),
    ("MELTONROSE", 29),
    ("EERSTE RIVER", 35),  # arrive
    ("EERSTE RIVER", 35),  # depart (same on source)
    ("FAURE", 38),
    ("FIRGROVE", 42),
    ("SOMERSET WEST", 54),
    ("VAN DER STEL", 57),
    ("STRAND", 60),
]

def ob_val(train, station_builder):
    return station_builder(train)

ob_rows = []
# Cape Town -> Bellville arrive
for name, off in ct_offs:
    row = [name]
    for tr in ob_trains:
        if tr in ct25:
            row.append(tadd(ct25[tr], off))
        else:
            row.append(E)
    ob_rows.append(row)

# PLATFORM NO
ob_rows.append(["PLATFORM NO"] + [platform_ob[tr] for tr in ob_trains])

# Bellville depart
ob_rows.append(["BELLVILLE"] + [bv_dep[tr] for tr in ob_trains])

# Strand branch stations (32xx only) interleaved in printed order with Kraaifontein at bottom
# Printed order after Bellville Dep: KUILS...STRAND then STIKLAND...KRAAIFONTEIN
for name, off in st_offs:
    row = [name]
    for tr in ob_trains:
        if tr.startswith("32"):
            row.append(tadd(bv_dep[tr], off))
        else:
            row.append(E)
    ob_rows.append(row)

for name, off in kf_offs:
    row = [name]
    for tr in ob_trains:
        if tr.startswith("25"):
            row.append(tadd(bv_dep[tr], off))
        else:
            row.append(E)
    ob_rows.append(row)

# Sanity outbound
assert ob_rows[0][1] == "06:15" and ob_rows[0][2] == E
assert ob_rows[8][1] == "06:48"  # VASCO
assert ob_rows[12][1] == "07:05"  # BELLVILLE arrive
assert ob_rows[13][1] == "4" and ob_rows[13][2] == "9"
assert ob_rows[14][1] == "07:06" and ob_rows[14][2] == "07:20"
assert ob_rows[15][1] == E and ob_rows[15][2] == "07:32"  # KUILS
assert ob_rows[-1][1] == "07:31" and ob_rows[-1][2] == E  # KRAAIFONTEIN
assert ob_rows[24][2] == "08:20"  # STRAND 3201

# ===== INBOUND =====
ib_trains = ["3200","2500","3202","2502","3204","2504","3206","2506","3208","2508","3210","2510","3212","2512"]

kf_start = {
    "2500": "06:38", "2502": "08:15", "2504": "09:40", "2506": "11:30",
    "2508": "13:00", "2510": "14:45", "2512": "16:00",
}
# offsets from Kraaifontein
kf_in_offs = [
    ("KRAAIFONTEIN", "D", 0),
    ("EIKENFONTEIN", "D", 5),
    ("BRACKENFELL", "D", 11),
    ("STIKLAND", "D", 18),  # 06:38+18=06:56
]

st_start = {
    "3200": "05:50", "3202": "07:15", "3204": "08:30", "3206": "10:45",
    "3208": "12:10", "3210": "14:00", "3212": "15:50",
}
st_in_offs = [
    ("STRAND", "D", 0),
    ("VAN DER STEL", "D", 3),
    ("SOMERSET WEST", "D", 6),
    ("FIRGROVE", "D", 18),
    ("FAURE", "D", 22),
    ("EERSTE RIVER", "A", 25),
    ("EERSTE RIVER", "D", 25),
    ("MELTONROSE", "D", 31),
    ("BLACKHEATH", "D", 38),
    ("KUILS RIVER", "D", 47),
]

bv_arr = {
    "3200": "06:50", "2500": "07:02", "3202": "08:15", "2502": "08:39",
    "3204": "09:30", "2504": "10:04", "3206": "11:45", "2506": "11:54",
    "3208": "13:10", "2508": "13:24", "3210": "15:00", "2510": "15:09",
    "3212": "16:50", "2512": "16:24",
}
platform_ib = {
    "3200": "9", "2500": "5", "3202": "9", "2502": "5",
    "3204": "9", "2504": "5", "3206": "3", "2506": "5",
    "3208": "3", "2508": "5", "3210": "9", "2510": "5",
    "3212": "9", "2512": "5",
}
bv_dep_ib = {
    "2500": "07:03", "2502": "08:40", "2504": "10:05", "2506": "11:55",
    "2508": "13:25", "2510": "15:10", "2512": "16:25",
}
# offsets from Bellville Dep to Cape Town (calibrated)
ct_in_offs = [
    ("TYGERBERG", "D", 6),
    ("PAROW", "D", 11),
    ("ELSIES RIVER", "D", 12),
    ("VASCO", "D", 16),
    ("GOODWOOD", "D", 19),
    ("THORNTON", "D", 24),
    ("MUTUAL", "D", 31),
    ("MAITLAND", "D", 35),
    ("KOEBERG RD", "D", 38),
    ("SALT RIVER", "D", 40),
    ("WOODSTOCK", "D", 44),
    ("CAPE TOWN", "A", 49),
]

ib_rows = []  # [station, ad, *times]

for name, ad, off in kf_in_offs:
    row = [name, ad]
    for tr in ib_trains:
        if tr in kf_start:
            row.append(tadd(kf_start[tr], off))
        else:
            row.append(E)
    ib_rows.append(row)

for name, ad, off in st_in_offs:
    row = [name, ad]
    for tr in ib_trains:
        if tr in st_start:
            row.append(tadd(st_start[tr], off))
        else:
            row.append(E)
    ib_rows.append(row)

ib_rows.append(["BELLVILLE", "A"] + [bv_arr[tr] for tr in ib_trains])
ib_rows.append(["PLATFORM NO", ""] + [platform_ib[tr] for tr in ib_trains])

row = ["BELLVILLE", "D"]
for tr in ib_trains:
    row.append(bv_dep_ib.get(tr, E))
ib_rows.append(row)

for name, ad, off in ct_in_offs:
    row = [name, ad]
    for tr in ib_trains:
        if tr in bv_dep_ib:
            row.append(tadd(bv_dep_ib[tr], off))
        else:
            row.append(E)
    ib_rows.append(row)

# Sanity inbound
assert ib_rows[0][3] == "06:38"  # KRAAIFONTEIN 2500 (cols: station, ad, 3200, 2500,...)
assert ib_rows[0][2] == E
assert ib_rows[3][3] == "06:56"  # STIKLAND 2500
assert ib_rows[4][2] == "05:50"  # STRAND 3200
assert ib_rows[14][2] == "06:50" and ib_rows[14][3] == "07:02"  # BELLVILLE A
assert ib_rows[-1][3] == "07:52"  # CAPE TOWN 2500
assert ib_rows[-1][2] == E
assert ib_rows[-1][15] == "17:14"  # CAPE TOWN 2512

# Styles
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
station_fill = PatternFill("solid", fgColor="D6EAF8")
station_font = Font(bold=True, name="Calibri", size=10)
title_font = Font(bold=True, name="Calibri", size=14, color="1F4E79")
subtitle_font = Font(bold=True, name="Calibri", size=11, color="1F4E79")
cell_font = Font(name="Calibri", size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
alt_fill = PatternFill("solid", fgColor="F5F9FC")
blank_fill = PatternFill("solid", fgColor="EEEEEE")
plat_fill = PatternFill("solid", fgColor="FFF2CC")

def style_cell(cell, is_station=False, is_blank=False, is_plat=False, alt=False):
    cell.border = thin
    if is_station:
        cell.font = station_font
        cell.fill = station_fill
        cell.alignment = left
    else:
        cell.font = cell_font
        cell.alignment = center
        if is_plat:
            cell.fill = plat_fill
        elif is_blank:
            cell.fill = blank_fill
        elif alt:
            cell.fill = alt_fill

wb = Workbook()

# Sheet 1 Outbound
ws1 = wb.active
ws1.title = "Outbound CT Bellville"
ws1["A1"] = "NORTHERN LINE PUBLIC HOLIDAY TIMETABLE 2026"
ws1["A1"].font = title_font
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ob_trains)+1)
ws1["A2"] = "Outbound: Cape Town / Bellville to Kraaifontein (25xx) and Strand (32xx)"
ws1["A2"].font = subtitle_font
ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ob_trains)+1)
ws1["A3"] = "Captured as-is from Metrorail / PRASA. '--' = train does not serve that station. 25xx: Cape Town→Kraaifontein. 32xx: start Bellville→Strand."
ws1["A3"].font = Font(italic=True, name="Calibri", size=9, color="666666")
ws1.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(ob_trains)+1)

headers = ["STATION / TRAIN NO."] + ob_trains
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=5, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin

for r_idx, row in enumerate(ob_rows):
    rn = 6 + r_idx
    is_plat = row[0] == "PLATFORM NO"
    for c_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=rn, column=c_idx, value=val)
        style_cell(cell, is_station=(c_idx==1), is_blank=(val==E), is_plat=is_plat, alt=(r_idx%2==1))

ws1.column_dimensions["A"].width = 18
for i in range(2, len(ob_trains)+2):
    ws1.column_dimensions[get_column_letter(i)].width = 7
ws1.freeze_panes = "B6"

# Sheet 2 Inbound
ws2 = wb.create_sheet("Inbound to Cape Town")
ws2["A1"] = "NORTHERN LINE PUBLIC HOLIDAY TIMETABLE 2026"
ws2["A1"].font = title_font
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ib_trains)+2)
ws2["A2"] = "Inbound: Kraaifontein (25xx) and Strand (32xx) to Bellville / Cape Town"
ws2["A2"].font = subtitle_font
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ib_trains)+2)
ws2["A3"] = "Captured as-is from Metrorail / PRASA. A=Arrive, D=Depart. 25xx continue Bellville→Cape Town. 32xx terminate at Bellville."
ws2["A3"].font = Font(italic=True, name="Calibri", size=9, color="666666")
ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(ib_trains)+2)

headers2 = ["STATION", "A/D"] + ib_trains
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=5, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin

for r_idx, row in enumerate(ib_rows):
    rn = 6 + r_idx
    is_plat = row[0] == "PLATFORM NO"
    for c_idx, val in enumerate(row, 1):
        cell = ws2.cell(row=rn, column=c_idx, value=val)
        style_cell(cell, is_station=(c_idx==1), is_blank=(val==E), is_plat=is_plat, alt=(r_idx%2==1))

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 6
for i in range(3, len(ib_trains)+3):
    ws2.column_dimensions[get_column_letter(i)].width = 7
ws2.freeze_panes = "C6"

# Sheet 3 long format
ws3 = wb.create_sheet("All trips (long format)")
for c, h in enumerate(["Direction", "Train No.", "Station", "A/D", "Station Order", "Time"], 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin

r = 2
for t_idx, tr in enumerate(ob_trains):
    for s_idx, row in enumerate(ob_rows):
        val = row[t_idx+1]
        if val == E:
            continue
        vals = ["Outbound", tr, row[0], "", s_idx+1, val]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            style_cell(cell, is_station=(c==3), alt=(r%2==0))
        r += 1

for t_idx, tr in enumerate(ib_trains):
    for s_idx, row in enumerate(ib_rows):
        val = row[t_idx+2]
        if val == E or val == "":
            continue
        vals = ["Inbound", tr, row[0], row[1], s_idx+1, val]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            style_cell(cell, is_station=(c==3), alt=(r%2==0))
        r += 1

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 6
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 10
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:F{r-1}"

wb.save(out)
print("Saved", out)
print("Outbound rows", len(ob_rows), "Inbound rows", len(ib_rows))
