from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timedelta

out = Path(r"C:\Users\enock\OneDrive\Documents\GitHub\Metrorail Next Train\Source Code\Southern-Line-Public-Holiday-Timetable-2026.xlsx")

def add_min(t, m):
    hh, mm = map(int, t.split(":"))
    dt = datetime(2000, 1, 1, hh, mm) + timedelta(minutes=m)
    return dt.strftime("%H:%M")

ob_trains = ["0101","0103","0105","0107","0109","0111","0113","0117","0119","0121","0123","0125","0127","0129","0131","0133","0135"]
ob_ct = ["05:30","06:10","06:50","07:20","08:10","08:40","10:00","11:40","12:20","13:00","14:00","14:30","15:10","15:40","16:00","16:50","17:35"]

ob_offsets = [
    ("CAPE TOWN", 0),
    ("WOODSTOCK", 3),
    ("SALT RIVER", 11),
    ("OBSERVATORY", 14),
    ("MOWBRAY", 16),
    ("ROSEBANK", 18),
    ("RONDEBOSCH", 20),
    ("NEWLANDS", 23),
    ("CLAREMONT", 25),
    ("HARFIELD ROAD", 27),
    ("KENILWORTH", 29),
    ("WYNBERG", 31),
    ("WITTEBOME", 33),
    ("PLUMSTEAD", 35),
    ("STEURHOF", 37),
    ("DIEP RIVER", 39),
    ("HEATHFIELD", 41),
    ("RETREAT", 43),  # arrive (bold on source)
    ("RETREAT", 44),  # depart
    ("STEENBERG", 46),
    ("LAKESIDE", 48),
    ("FALSE BAY", 51),
    ("MUIZENBERG", 53),
    ("ST. JAMES", 56),
    ("KALK BAY", 58),
    ("FISH HOEK", 63),  # arrive
]

ob_continue = {"0101","0105","0109","0113","0121","0125","0131"}
ob_simon_offsets = [
    ("FISH HOEK", 0),  # depart / continue
    ("SUNNY COVE", 5),
    ("GLENCAIRN", 14),
    ("SIMONSTOWN", 20),
]

ob_rows = []
for name, off in ob_offsets:
    ob_rows.append([name] + [add_min(t, off) for t in ob_ct])

fh_arrive = [add_min(t, 63) for t in ob_ct]
for name, off in ob_simon_offsets:
    row = [name]
    for i, train in enumerate(ob_trains):
        if train in ob_continue:
            row.append(add_min(fh_arrive[i], off))
        else:
            row.append("..")
    ob_rows.append(row)

ib_trains = ["0100","0102","0104","0106","0108","0110","0112","0114","0116","0118","0120","0122","0124","0126","0128","0130","0132"]
ib_simon = {
    "0102": "05:30",
    "0106": "07:10",
    "0112": "08:40",
    "0116": "10:10",
    "0118": "11:33",
    "0124": "14:33",
    "0128": "16:05",
}
ib_simon_offsets = [
    ("SIMON'S TOWN", 0),
    ("GLENCAIRN", 6),
    ("SUNNY COVE", 15),
    ("FISH HOEK", 20),  # arrive from Simonstown
]
ib_fh = ["05:30","05:50","06:40","07:30","07:50","08:30","09:00","10:00","10:30","11:53","13:00","14:00","14:53","15:30","16:25","16:50","17:10"]
ib_offsets = [
    ("FISH HOEK", 0),  # depart
    ("KALK BAY", 4),
    ("ST. JAMES", 6),
    ("MUIZENBERG", 9),
    ("FALSE BAY", 11),
    ("LAKESIDE", 13),
    ("STEENBERG", 15),
    ("RETREAT", 17),  # arrive
    ("RETREAT", 18),  # depart
    ("HEATHFIELD", 21),
    ("DIEP RIVER", 24),
    ("STEURHOF", 27),
    ("PLUMSTEAD", 29),
    ("WITTEBOME", 31),
    ("WYNBERG", 33),
    ("KENILWORTH", 35),
    ("HARFIELD ROAD", 37),
    ("CLAREMONT", 39),
    ("NEWLANDS", 41),
    ("RONDEBOSCH", 43),
    ("ROSEBANK", 45),
    ("MOWBRAY", 47),
    ("OBSERVATORY", 49),
    ("SALT RIVER", 52),
    ("WOODSTOCK", 60),
    ("CAPE TOWN", 63),
]

ib_rows = []
for name, off in ib_simon_offsets:
    row = [name]
    for train in ib_trains:
        if train in ib_simon:
            row.append(add_min(ib_simon[train], off))
        else:
            row.append("..")
    ib_rows.append(row)
for name, off in ib_offsets:
    ib_rows.append([name] + [add_min(t, off) for t in ib_fh])

# Anchors
assert ob_rows[0][1] == "05:30"
assert ob_rows[25][1] == "06:33"  # FISH HOEK arrive
assert ob_rows[26][1] == "06:33" and ob_rows[26][2] == ".."
assert ob_rows[29][1] == "06:53" and ob_rows[29][10] == "14:23"
assert ib_rows[0][2] == "05:30" and ib_rows[0][1] == ".."
assert ib_rows[4][1] == "05:30"  # FISH HOEK depart 0100
assert ib_rows[-1][1] == "06:33" and ib_rows[-1][10] == "12:56"

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
station_fill = PatternFill("solid", fgColor="D6EAF8")
station_font = Font(bold=True, name="Calibri", size=10)
title_font = Font(bold=True, name="Calibri", size=14, color="1F4E79")
subtitle_font = Font(bold=True, name="Calibri", size=12, color="1F4E79")
cell_font = Font(name="Calibri", size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
alt_fill = PatternFill("solid", fgColor="F5F9FC")
blank_fill = PatternFill("solid", fgColor="EEEEEE")

def write_sheet(ws, title, direction, trains, rows):
    ws["A1"] = title
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(trains)+1)
    ws["A2"] = direction
    ws["A2"].font = subtitle_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(trains)+1)
    ws["A3"] = "Captured as-is from Metrorail / PRASA Southern Line Public Holiday Timetable 2026. '..' means the train does not serve that station."
    ws["A3"].font = Font(italic=True, name="Calibri", size=9, color="666666")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(trains)+1)
    headers = ["STATION / TRAIN NO."] + trains
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    for r_idx, row in enumerate(rows):
        row_num = 6 + r_idx
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.border = thin
            if c_idx == 1:
                cell.font = station_font
                cell.fill = station_fill
                cell.alignment = left
            else:
                cell.font = cell_font
                cell.alignment = center
                if val == "..":
                    cell.fill = blank_fill
                elif r_idx % 2 == 1:
                    cell.fill = alt_fill
    ws.column_dimensions["A"].width = 18
    for i in range(2, len(trains)+2):
        ws.column_dimensions[get_column_letter(i)].width = 7
    ws.freeze_panes = "B6"

wb = Workbook()
ws1 = wb.active
ws1.title = "Cape Town to Simons Town"
write_sheet(ws1, "SOUTHERN LINE PUBLIC HOLIDAY TIMETABLE 2026", "Cape Town to Simon's Town", ob_trains, ob_rows)
ws2 = wb.create_sheet("Simons Town to Cape Town")
write_sheet(ws2, "SOUTHERN LINE PUBLIC HOLIDAY TIMETABLE 2026", "Simon's Town to Cape Town", ib_trains, ib_rows)

ws3 = wb.create_sheet("All trips (long format)")
for col, h in enumerate(["Direction", "Train No.", "Station", "Station Order", "Time"], 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin
row = 2
for direction, trains, rows in [
    ("Cape Town to Simon's Town", ob_trains, ob_rows),
    ("Simon's Town to Cape Town", ib_trains, ib_rows),
]:
    for t_idx, train in enumerate(trains):
        for s_idx, r in enumerate(rows):
            val = r[t_idx + 1]
            if val == "..":
                continue
            for c, v in enumerate([direction, train, r[0], s_idx + 1, val], 1):
                cell = ws3.cell(row=row, column=c, value=v)
                cell.border = thin
                cell.font = cell_font
                cell.alignment = center if c != 3 else left
            row += 1
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 10
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = f"A1:E{row-1}"

wb.save(out)

def write_tsv(path, trains, rows, title, direction):
    lines = [title, direction, "Captured as-is from Metrorail / PRASA Southern Line Public Holiday Timetable 2026. '..' = no service.", "", "\t".join(["STATION / TRAIN NO."] + trains)]
    for r in rows:
        lines.append("\t".join(r))
    path.write_text("\n".join(lines), encoding="utf-8")

write_tsv(out.parent / "southern_outbound.tsv", ob_trains, ob_rows, "SOUTHERN LINE PUBLIC HOLIDAY TIMETABLE 2026", "Cape Town to Simon's Town")
write_tsv(out.parent / "southern_inbound.tsv", ib_trains, ib_rows, "SOUTHERN LINE PUBLIC HOLIDAY TIMETABLE 2026", "Simon's Town to Cape Town")
print("OK", out)
